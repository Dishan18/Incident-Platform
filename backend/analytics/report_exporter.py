"""
Report Exporter Module.
Supports exporting custom SQL analytics query results to Excel, CSV, and PDF formats.
"""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def export_to_csv(df: pd.DataFrame) -> bytes:
    """Export the dataframe directly to CSV bytes."""
    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode("utf-8")


def export_to_excel(prompt: str, sql_query: str, summary_text: str, df: pd.DataFrame) -> bytes:
    """Generate a professionally styled Excel spreadsheet of query results."""
    wb = openpyxl.Workbook()
    # Remove default worksheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    font_family = "Segoe UI"
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    title_font = Font(name=font_family, size=14, bold=True, color="1F2937")

    border_color = "D1D5DB"  # tailwind gray-300
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )

    # ── Sheet 1: Summary Sheet ──
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.append([])
    ws_summary.append(["AI Custom Analytics Report"])
    ws_summary.cell(row=2, column=1).font = title_font
    ws_summary.append([])

    # Metrics Table
    summary_data = [
        ("Natural Language Request", prompt),
        ("Execution Time (Local)", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Rows Found", len(df)),
    ]

    for label, val in summary_data:
        ws_summary.append([label, val])
        r = ws_summary.max_row
        ws_summary.cell(row=r, column=1).font = bold_font
        ws_summary.cell(row=r, column=2).font = regular_font
        ws_summary.cell(row=r, column=1).border = thin_border
        ws_summary.cell(row=r, column=2).border = thin_border

    ws_summary.append([])
    ws_summary.append(["Query Results Summary"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font

    if summary_text:
        for line in summary_text.split("\n"):
            if line.strip():
                ws_summary.append([line.strip()])
                ws_summary.cell(row=ws_summary.max_row, column=1).font = regular_font

    # Autofit Column Widths
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > 120:
                    max_len = max(max_len, 60)
                else:
                    max_len = max(max_len, len(val_str))
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 22)

    # ── Sheet 2: Query Results ──
    ws_results = wb.create_sheet(title="Query Results")
    ws_results.views.sheetView[0].showGridLines = True

    # Very minimal heading
    ws_results.append(["Custom Query Results"])
    ws_results.cell(row=1, column=1).font = Font(name=font_family, size=13, bold=True, color="1E293B")

    ws_results.append([f"Request: {prompt}"])
    ws_results.cell(row=2, column=1).font = Font(name=font_family, size=10, italic=True)

    ws_results.append([f"Rows Recovered: {len(df)}"])
    ws_results.cell(row=3, column=1).font = Font(name=font_family, size=10)

    # Blank spacer row
    ws_results.append([])

    # Header Row (Row 5)
    headers = list(df.columns)
    ws_results.append(headers)

    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate slate-800
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # slate-50

    for col_idx in range(1, len(headers) + 1):
        cell = ws_results.cell(row=5, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows (Row 6 onwards)
    start_row = 6
    for r_offset, row in enumerate(df.itertuples(index=False)):
        r_idx = start_row + r_offset
        row_vals = []
        for val in row:
            if pd.isna(val):
                row_vals.append(None)
            elif isinstance(val, (pd.Timestamp, pd.DatetimeIndex)):
                row_vals.append(val.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                row_vals.append(val)
        ws_results.append(row_vals)

        row_fill = zebra_fill if r_idx % 2 == 0 else None
        for c_idx in range(1, len(headers) + 1):
            cell = ws_results.cell(row=r_idx, column=c_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    # Freeze Header row at row 6
    ws_results.freeze_panes = "A6"

    # Add Data filters starting on row 5
    ws_results.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{len(df) + 5}"

    # Auto column widths
    for col in ws_results.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > 100:
                    max_len = max(max_len, 40)
                else:
                    max_len = max(max_len, len(val_str))
        ws_results.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # ── Sheet 3: SQL Script ──
    ws_sql = wb.create_sheet(title="SQL Script")
    ws_sql.views.sheetView[0].showGridLines = True
    ws_sql.append(["SQL Query Used"])
    ws_sql.cell(row=1, column=1).font = bold_font
    ws_sql.append([])

    for line in sql_query.split("\n"):
        ws_sql.append([line])
        ws_sql.cell(row=ws_sql.max_row, column=1).font = Font(name="Consolas", size=10, color="2563EB")  # blue-600

    ws_sql.column_dimensions["A"].width = 90

    # Set the Results sheet as active so it opens first
    wb.active = ws_results

    # Save workbook to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_to_pdf(prompt: str, sql_query: str, summary_text: str, chart_rec: dict, df: pd.DataFrame) -> bytes:
    """Generate a clean, high-end PDF report of the custom analytics queries."""
    buffer = io.BytesIO()

    # Define standard margins and layout (Landscape letter fits wide columns nicely)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=45,
        bottomMargin=45
    )

    styles = getSampleStyleSheet()

    # Color Tokens
    primary_color = colors.HexColor('#0F172A')   # Slate 900
    accent_color = colors.HexColor('#2563EB')    # Blue 600
    border_color = colors.HexColor('#E2E8F0')    # Slate 200
    bg_light = colors.HexColor('#F8FAFC')        # Slate 50
    text_color = colors.HexColor('#334155')      # Slate 700

    # Paragraph Styles
    title_style = ParagraphStyle(
        'PDFTitle',
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=primary_color,
        spaceAfter=4
    )

    subtitle_style = ParagraphStyle(
        'PDFSub',
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )

    section_heading = ParagraphStyle(
        'PDFSection',
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=16,
        textColor=primary_color,
        spaceBefore=12,
        spaceAfter=6,
        keepWithNext=True
    )

    body_style = ParagraphStyle(
        'PDFBody',
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=text_color,
        spaceAfter=8
    )

    code_style = ParagraphStyle(
        'PDFCode',
        fontName='Courier',
        fontSize=8.5,
        leading=11,
        textColor=accent_color,
        spaceAfter=10
    )

    table_header_style = ParagraphStyle(
        'PDFTableHeader',
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9,
        textColor=colors.white
    )

    table_cell_style = ParagraphStyle(
        'PDFTableCell',
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        textColor=text_color
    )

    story = []

    # Title & Metadata
    story.append(Paragraph("AI Custom Analytics Report", title_style))
    gen_time = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated On: {gen_time} | Report Scope: DuckDB read-only query engine", subtitle_style))

    # Details Block Table
    details_data = [
        [Paragraph("<b>Natural Language Request:</b>", body_style), Paragraph(prompt, body_style)],
        [Paragraph("<b>Total Rows Returned:</b>", body_style), Paragraph(f"{len(df)} rows", body_style)]
    ]
    details_table = Table(details_data, colWidths=[150, 570])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), bg_light),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 0.5, border_color),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, border_color),
    ]))
    story.append(details_table)
    story.append(Spacer(1, 10))

    # SQL query code block
    story.append(Paragraph("Executed SQL Code", section_heading))
    sql_lines = "<br/>".join(sql_query.replace("\n", "<br/>").split("<br/>"))
    sql_box = Table([[Paragraph(sql_lines, code_style)]], colWidths=[720])
    sql_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F5F9')),
        ('BOX', (0, 0), (-1, -1), 0.5, border_color),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(sql_box)
    story.append(Spacer(1, 10))

    # Executive summary
    if summary_text:
        story.append(Paragraph("Query Results Summary", section_heading))
        summary_html = "<br/>".join([f"• {line.strip('- ').strip()}" if line.strip().startswith("-") else line.strip() 
                                     for line in summary_text.split("\n") if line.strip()])
        story.append(Paragraph(summary_html, body_style))
        story.append(Spacer(1, 10))

    # Recommended Chart Metadata (if any)
    if chart_rec and chart_rec.get("chart") not in [None, "none"]:
        story.append(Paragraph("Recommended Visualization Details", section_heading))
        chart_details = (
            f"<b>Recommended Chart:</b> {chart_rec.get('chart').upper()}<br/>"
            f"<b>Title:</b> {chart_rec.get('title', 'N/A')}<br/>"
            f"<b>X-Axis Mapping:</b> {chart_rec.get('x', 'N/A')} | <b>Y-Axis Mapping:</b> {chart_rec.get('y', 'N/A')}"
        )
        story.append(Paragraph(chart_details, body_style))
        story.append(Spacer(1, 10))

    # Data Table Summary (limit table to first 100 rows to prevent bloated PDFs)
    story.append(Paragraph("Query Results Table", section_heading))
    limit_rows = 100
    df_to_render = df.head(limit_rows)

    headers = list(df.columns)
    table_data = []
    # Header cells
    table_data.append([Paragraph(h, table_header_style) for h in headers])

    # Row cells
    for row in df_to_render.itertuples(index=False):
        row_cells = []
        for val in row:
            if pd.isna(val):
                val_str = ""
            elif isinstance(val, (pd.Timestamp, pd.DatetimeIndex)):
                val_str = val.strftime("%Y-%m-%d %H:%M:%S")
            else:
                val_str = str(val)
            row_cells.append(Paragraph(val_str, table_cell_style))
        table_data.append(row_cells)

    # Dynamic column widths calculation to fit landscape width (720pt printable area)
    num_cols = len(headers)
    col_w = 720.0 / num_cols
    col_widths = [col_w] * num_cols

    results_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BOX', (0, 0), (-1, -1), 0.5, border_color),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, border_color),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, bg_light])
    ]))
    story.append(results_table)

    if len(df) > limit_rows:
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"<i>* Showing first {limit_rows} rows. Remaining {len(df) - limit_rows} rows have been exported in the corresponding Excel spreadsheet.</i>", body_style))

    # Footer layout callback
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7.5)
        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.line(36, 32, 756, 32)
        canvas.drawString(36, 20, "Confidential - Operational Incident Management Command Center")
        canvas.drawRightString(756, 20, f"Page {doc.page} | Generated: {gen_time}")
        canvas.restoreState()

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes
